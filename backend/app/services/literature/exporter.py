from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.config import SCREENING_COLUMNS


def _split_field(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = [part.strip() for part in text.split(";")]
    return [part for part in parts if part]


def dataframe_to_csv_bytes(dataframe: pd.DataFrame) -> bytes:
    buffer = io.StringIO()
    dataframe.to_csv(buffer, index=False)
    return buffer.getvalue().encode("utf-8-sig")


def dataframe_to_excel_bytes(dataframe: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
    return buffer.getvalue()


def dataframe_to_ris_bytes(dataframe: pd.DataFrame) -> bytes:
    """Write a best-effort RIS export preserving standard metadata fields."""
    lines: list[str] = []
    for _, row in dataframe.fillna("").iterrows():
        authors = _split_field(row.get("authors", ""))
        keywords = _split_field(row.get("keywords", ""))
        author_lines = [f"AU  - {author}" for author in authors] or ["AU  - "]
        keyword_lines = [f"KW  - {keyword}" for keyword in keywords] or ["KW  - "]
        lines.extend(
            [
                "TY  - JOUR",
                f"TI  - {row.get('title', '')}",
                *author_lines,
                f"PY  - {row.get('year', '')}",
                f"JO  - {row.get('journal', '')}",
                f"DO  - {row.get('doi', '')}",
                f"AB  - {row.get('abstract', '')}",
                *keyword_lines,
                f"ID  - {row.get('pmid', '')}",
                f"DB  - {row.get('source_database', '')}",
                "ER  - ",
                "",
            ]
        )
    return "\n".join(lines).encode("utf-8-sig")


def dataframe_to_nbib_bytes(dataframe: pd.DataFrame) -> bytes:
    """Write a PubMed-style NBIB export for downstream review workflows."""
    lines: list[str] = []
    for _, row in dataframe.fillna("").iterrows():
        authors = _split_field(row.get("authors", ""))
        keywords = _split_field(row.get("keywords", ""))
        author_lines = [f"AU  - {author}" for author in authors] or ["AU  - "]
        keyword_lines = [f"OT  - {keyword}" for keyword in keywords] or ["OT  - "]
        lines.extend(
            [
                f"PMID- {row.get('pmid', '')}",
                f"TI  - {row.get('title', '')}",
                *author_lines,
                f"DP  - {row.get('year', '')}",
                f"JT  - {row.get('journal', '')}",
                f"AID - {row.get('doi', '')} [doi]",
                f"AB  - {row.get('abstract', '')}",
                f"LID - {row.get('doi', '')} [doi]",
                *keyword_lines,
                f"DB  - {row.get('source_database', '')}",
                "",
            ]
        )
    return "\n".join(lines).encode("utf-8-sig")


def create_screening_excel(dataframe: pd.DataFrame, review_dataset: pd.DataFrame | None = None) -> bytes:
    """Create an Excel workbook directly usable for screening.

    The first sheet contains the screening-ready table. A second sheet records
    duplicate-review decisions when available.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Screening"

    screening = pd.DataFrame(columns=SCREENING_COLUMNS)
    source = dataframe.copy()
    screening["Title"] = source.get("title", "")
    screening["Authors"] = source.get("authors", "")
    screening["Year"] = source.get("year", "")
    screening["Journal"] = source.get("journal", "")
    screening["DOI"] = source.get("doi", "")
    screening["PMID"] = source.get("pmid", "")
    screening["Abstract"] = source.get("abstract", "")
    screening["Keywords"] = source.get("keywords", "")
    screening["Database Source"] = source.get("source_database", "")
    screening["Include"] = ""
    screening["Exclude"] = ""
    screening["Maybe"] = ""
    screening["Reason for Exclusion"] = ""
    screening["Reviewer"] = ""
    screening["Notes"] = ""

    for row in screening.itertuples(index=False):
        sheet.append(list(row))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    if review_dataset is not None and not review_dataset.empty:
        review_sheet = workbook.create_sheet("Duplicate Review")
        for row in review_dataset.itertuples(index=False):
            review_sheet.append(list(row))
        for cell in review_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_report_text(report: dict[str, Any]) -> str:
    lines = ["SODH Literature Toolkit Processing Report", ""]
    for key, value in report.items():
        lines.append(f"{key}: {value}")
    return "\n".join(lines)
